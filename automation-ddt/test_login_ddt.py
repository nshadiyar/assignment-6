"""
Assignment 6: Data-Driven and Cross-Browser Automated Testing
Login tests using Excel data on Sauce Labs
"""

from pathlib import Path
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.firefox.options import Options as FirefoxOptions
import os

# ================== CONFIG ==================

BASE_URL = "https://practicetestautomation.com/practice-test-login/"
TEST_DATA_FILE = "test_data.xlsx"
SHEET_NAME = "Sheet1"
TIMEOUT = 10

USERNAME_INPUT_ID = "username"
PASSWORD_INPUT_ID = "password"
SUBMIT_BUTTON_ID = "submit"

SUCCESS_TEXT = "Logged In Successfully"
ERROR_TEXT = "invalid"

# ================== DATA ==================

def load_test_data():
    path = Path(__file__).parent / TEST_DATA_FILE
    if not path.exists():
        raise FileNotFoundError("test_data.xlsx not found")

    wb = load_workbook(path)
    sheet = wb[SHEET_NAME]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        data.append({
            "username": str(row[0]).strip(),
            "password": str(row[1]).strip(),
            "expected": str(row[2]).strip().upper()
        })
    wb.close()
    return data

# ================== DRIVER ==================

def create_sauce_driver(browser_name, test_name):
    username = os.getenv("SAUCE_USERNAME")
    access_key = os.getenv("SAUCE_ACCESS_KEY")
    if not username or not access_key:
        raise ValueError("SAUCE_USERNAME and SAUCE_ACCESS_KEY must be set")

    sauce_options = {
        "name": test_name,
        "build": "Assignment6-DDT",
    }

    if browser_name.lower() == "firefox":
        options = FirefoxOptions()
        options.set_capability("browserVersion", "latest")
    else:
        options = ChromeOptions()
        options.set_capability("browserVersion", "latest")

    # Let Sauce Labs pick a supported platform (e.g. Windows 10/11)
    options.set_capability("platformName", "Windows 11")
    options.set_capability("sauce:options", sauce_options)

    # Region: US West (default) or EU Central if SAUCE_ENDPOINT=eu
    region = "eu-central-1" if os.getenv("SAUCE_ENDPOINT", "").strip().lower() == "eu" else "us-west-1"
    base = f"https://ondemand.{region}.saucelabs.com/wd/hub"
    hub_url = f"https://{username}:{access_key}@{base.split('://')[1]}"
    driver = webdriver.Remote(command_executor=hub_url, options=options)
    return driver

# ================== TEST ==================

def run_login_test(driver, username, password, expected):
    driver.get(BASE_URL)
    wait = WebDriverWait(driver, TIMEOUT)

    wait.until(EC.presence_of_element_located((By.ID, USERNAME_INPUT_ID))).send_keys(username)
    driver.find_element(By.ID, PASSWORD_INPUT_ID).send_keys(password)
    driver.find_element(By.ID, SUBMIT_BUTTON_ID).click()

    try:
        wait.until(lambda d: SUCCESS_TEXT.lower() in d.page_source.lower()
                            or ERROR_TEXT in d.page_source.lower())
    except TimeoutException:
        return False, "Timeout"

    page = driver.page_source.lower()
    actual = "SUCCESS" if SUCCESS_TEXT.lower() in page else "ERROR"
    return actual == expected, f"expected={expected}, actual={actual}"

# ================== RUNNER ==================

def main():
    # Кросс-браузер тесты на Sauce Labs
    browsers = ["chrome", "firefox"]
    test_data = load_test_data()

    for browser in browsers:
        print(f"\nRunning tests | Sauce Labs=True | Browser={browser}")
        print("-" * 50)

        for i, row in enumerate(test_data, start=1):
            driver = None
            try:
                test_name = f"Login DDT - Row {i} ({row['expected']}) - {browser}"
                driver = create_sauce_driver(browser, test_name)
                passed, msg = run_login_test(driver, row["username"], row["password"], row["expected"])
                print(f"Dataset {i}: {'PASSED' if passed else 'FAILED'} - {msg}")
            finally:
                if driver:
                    driver.quit()

if __name__ == "__main__":
    main()